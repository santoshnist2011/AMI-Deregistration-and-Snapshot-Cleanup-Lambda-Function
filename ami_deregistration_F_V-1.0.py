import boto3
import os
from datetime import datetime, timedelta
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Alignment

def lambda_handler(event, context):
    # Initialize the SES client (use the same region as the Lambda function)
    ses_client = boto3.client('ses', region_name=os.environ['AWS_REGION'])

    # Get regions from environment variable
    regions = os.environ['AWS_REGIONS'].split(',')

    # Initialize a list to store details for the email
    ami_details = []

    for region in regions:
        ec2_client = boto3.client('ec2', region_name=region)

        # Fetch AMIs based on tags
        images_response = ec2_client.describe_images(
            Filters=[
                {'Name': 'tag:Purpose', 'Values': ['Patching']},
                {'Name': 'tag:Delete', 'Values': ['Yes']}
            ],
            Owners=['self']
        )

        for image in images_response['Images']:
            ami_id = image['ImageId']
            creation_date = image['CreationDate']
            creation_date_dt = datetime.strptime(creation_date, '%Y-%m-%dT%H:%M:%S.%fZ')

            # Check if the AMI is older than one week
            if datetime.now() - creation_date_dt > timedelta(days=7):
                snapshots = [
                    block_device.get('Ebs', {}).get('SnapshotId')
                    for block_device in image.get('BlockDeviceMappings', [])
                    if block_device.get('Ebs', {}).get('SnapshotId')
                ]

                # Deregister AMI and delete associated snapshots
                ec2_client.deregister_image(ImageId=ami_id)
                for snapshot_id in snapshots:
                    ec2_client.delete_snapshot(SnapshotId=snapshot_id)

                # Append details for the Excel report
                ami_details.append({
                    'Account': image.get('OwnerId', 'N/A'),
                    'Region': region,
                    'AMI ID': ami_id,
                    'Snapshots': snapshots,
                    'Delete Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })

    # Get sender and recipient email addresses from environment variables
    sender_email = os.environ['SENDER_EMAIL']
    recipient_emails = os.environ['RECIPIENT_EMAILS'].split(',')

    # Verify email addresses
    verified_emails = get_verified_emails(ses_client)
    valid_recipient_emails = [email for email in recipient_emails if email in verified_emails]
    unverified_emails = [email for email in recipient_emails if email not in verified_emails]

    if not valid_recipient_emails:
        return {
            'statusCode': 400,
            'body': 'No valid recipient email addresses found.'
        }

    # Send email with SES
    try:
        if ami_details:
            # Create an Excel workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "AMI Snapshot Report"

            # Add headers
            headers = ['Account', 'Region', 'AMI ID', 'Snapshot', 'Delete Date']
            ws.append(headers)

            # Add data to worksheet
            for detail in ami_details:
                snapshots = detail['Snapshots']
                for i, snapshot in enumerate(snapshots):
                    row = [
                        detail['Account'],
                        detail['Region'],
                        detail['AMI ID'] if i == 0 else "",
                        snapshot,
                        detail['Delete Date'] if i == 0 else ""
                    ]
                    ws.append(row)

                # Merge cells for AMI ID and Delete Date for multiple snapshots
                if len(snapshots) > 1:
                    ws.merge_cells(start_row=ws.max_row - len(snapshots) + 1, end_row=ws.max_row, start_column=3, end_column=3)
                    ws.merge_cells(start_row=ws.max_row - len(snapshots) + 1, end_row=ws.max_row, start_column=5, end_column=5)
                    # Align the merged cells
                    for col in [3, 5]:
                        cell = ws.cell(row=ws.max_row - len(snapshots) + 1, column=col)
                        cell.alignment = Alignment(vertical='center')

            # Save the workbook to a temporary file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                excel_file_path = tmp.name
            wb.save(excel_file_path)

            # Read the Excel file content
            with open(excel_file_path, 'rb') as f:
                excel_data = f.read()

            # Send email with Excel attachment
            response = ses_client.send_raw_email(
                Source=sender_email,
                Destinations=valid_recipient_emails,
                RawMessage={
                    'Data': create_raw_email_with_attachment(sender_email, valid_recipient_emails, excel_data, unverified_emails)
                }
            )
        else:
            # Send email indicating no AMIs older than 7 days were found
            response = ses_client.send_raw_email(
                Source=sender_email,
                Destinations=valid_recipient_emails,
                RawMessage={
                    'Data': create_raw_email_no_ami_found(sender_email, valid_recipient_emails, unverified_emails)
                }
            )
    except ses_client.exceptions.MessageRejected as e:
        return {
            'statusCode': 400,
            'body': f'Failed to send email: {str(e)}'
        }

    return {
        'statusCode': 200,
        'body': 'Successfully processed AMIs and sent email.'
    }

def get_verified_emails(ses_client):
    """Fetch a list of verified email addresses from SES."""
    response = ses_client.list_identities(IdentityType='EmailAddress')
    verified_emails = []
    for identity in response['Identities']:
        verification_attributes = ses_client.get_identity_verification_attributes(Identities=[identity])
        if verification_attributes['VerificationAttributes'].get(identity, {}).get('VerificationStatus') == 'Success':
            verified_emails.append(identity)
    return verified_emails

def create_raw_email_with_attachment(source_email, dest_emails, excel_data, unverified_emails):
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication

    msg = MIMEMultipart()
    msg['Subject'] = 'AMI and Snapshot Deregistration Report'
    msg['From'] = source_email
    msg['To'] = ', '.join(dest_emails)

    # Create a custom email body
    email_body = f"""
    Hi Team,

    Good day! Hope you are doing well.
    We are pleased to inform you that the manually created AMIs older than one week and their associated
    snapshots have been successfully deleted.

    The attached Excel file contains details of the deregistered AMIs and deleted
    snapshots.

    Unverified email addresses that did not receive this report:
    {', '.join(unverified_emails) if unverified_emails else 'None'}

    ***This is an auto generated Email. Do not reply to this email.***

    Best regards,
    NTT Data INC.
    """

    # Add body to the email
    body = MIMEText(email_body, 'plain')
    msg.attach(body)

    # Attach the Excel file
    attachment = MIMEApplication(excel_data, _subtype='xlsx')
    attachment.add_header('Content-Disposition', 'attachment', filename='ami_snapshot_report.xlsx')
    msg.attach(attachment)

    return msg.as_string()

def create_raw_email_no_ami_found(source_email, dest_emails, unverified_emails):
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    msg = MIMEMultipart()
    msg['Subject'] = 'No AMIs Older Than 7 Days Found'
    msg['From'] = source_email
    msg['To'] = ', '.join(dest_emails)

    # Create a custom email body for no AMI found
    email_body = f"""
    Hi Team,

    Good day! Hope you are doing well.
    We checked for manually created AMIs older than one week, but none were found.

    Unverified email addresses that did not receive this report:
    {', '.join(unverified_emails) if unverified_emails else 'None'}

    ***This is an auto generated Email. Do not reply to this email.***

    Best regards,
    NTT Data INC.
    """

    # Add body to the email
    body = MIMEText(email_body, 'plain')
    msg.attach(body)

    return msg.as_string()
